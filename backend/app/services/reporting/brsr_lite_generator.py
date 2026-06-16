"""
services/reporting/brsr_lite_generator.py
------------------------------------------
Generates BRSR Lite Excel report for an MSME.

Sheets:
  1. Cover          — Company info, FY, generated date
  2. Energy         — Scope 1 + 2, monthly breakdown, intensity
  3. Water          — Withdrawal, consumption, recycled, intensity
  4. Emissions      — Scope 1 + 2 summary with source breakdown
  5. Water Quality  — Lab results and compliance status per parameter
  6. Data Quality   — Actual vs estimated, gaps, confidence score
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette — SANGATNA brand
# ---------------------------------------------------------------------------
C_GREEN_DARK   = "1A6B45"
C_GREEN_MID    = "2D9E6B"
C_GREEN_LIGHT  = "E8F5EE"
C_GRAY_DARK    = "2C2C2C"
C_GRAY_MID     = "6B7280"
C_GRAY_LIGHT   = "F3F4F6"
C_WHITE        = "FFFFFF"
C_AMBER        = "F59E0B"
C_AMBER_LIGHT  = "FEF3C7"
C_RED          = "DC2626"
C_RED_LIGHT    = "FEE2E2"
C_BLUE_LIGHT   = "EFF6FF"
C_BLUE         = "1D4ED8"
C_PURPLE_LIGHT = "F5F3FF"
C_PURPLE       = "7C3AED"


def _font(bold=False, size=10, color=C_GRAY_DARK, italic=False):
    return Font(
        name="Arial", bold=bold, size=size,
        color=color, italic=italic
    )


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _border(style="thin"):
    s = Side(style=style, color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right():
    return Alignment(horizontal="right", vertical="center")


def _header_cell(ws, row, col, value, bg=C_GREEN_DARK, fg=C_WHITE,
                 bold=True, size=10, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, color=fg)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _center() if align == "center" else _left()
    return cell


def _data_cell(ws, row, col, value, bg=C_WHITE, bold=False,
               align="left", number_format=None, color=C_GRAY_DARK):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _right() if align == "right" else _left()
    if number_format:
        cell.number_format = number_format
    return cell


def _section_title(ws, row, col, value, width=8):
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row,   end_column=col + width - 1
    )
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=True, size=11, color=C_WHITE)
    cell.fill = _fill(C_GREEN_MID)
    cell.alignment = _left()
    cell.border = _border()
    return cell


def _status_cell(ws, row, col, status):
    """Renders SAFE/EXCEEDS/ESTIMATED with colour coding."""
    colors = {
        "SAFE":          (C_GREEN_LIGHT, C_GREEN_DARK),
        "EXCEEDS_LIMIT": (C_RED_LIGHT,   C_RED),
        "NO_LIMIT":      (C_GRAY_LIGHT,  C_GRAY_MID),
        "NO_DATA":       (C_AMBER_LIGHT, C_AMBER),
        "ESTIMATED":     (C_AMBER_LIGHT, C_AMBER),
        "ACTUAL":        (C_GREEN_LIGHT, C_GREEN_DARK),
        "MISSING":       (C_RED_LIGHT,   C_RED),
    }
    bg, fg = colors.get(status, (C_WHITE, C_GRAY_DARK))
    cell = ws.cell(row=row, column=col, value=status.replace("_", " "))
    cell.font = _font(bold=True, color=fg)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _center()
    return cell


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_brsr_lite(data: dict) -> bytes:
    """
    data dict structure:
    {
      "company": { name, industry, city, state, gstin, financial_year },
      "facilities": [ { id, name, state, city } ],
      "energy": {
        "scope1_co2e_tonnes": float,
        "scope2_co2e_tonnes": float,
        "total_co2e_tonnes":  float,
        "total_kwh":          float,
        "total_diesel_litres":float,
        "monthly": [
          { period, facility_name, source, kwh, diesel_litres,
            co2e_tonnes, is_estimated }
        ]
      },
      "water": {
        "total_withdrawal_kl":  float,
        "total_consumption_kl": float,
        "total_recycled_kl":    float,
        "total_discharged_kl":  float,
        "monthly": [
          { period, facility_name, source, category, quantity_kl,
            is_estimated }
        ]
      },
      "water_quality": [
        {
          "lab_name": str,
          "collection_date": str,
          "water_type": str,
          "overall_status": str,
          "readings": [
            { parameter, measured_value, unit, status, bis_limit, cpcb_limit }
          ]
        }
      ],
      "emissions": {
        "scope1_co2e_tonnes": float,
        "scope2_co2e_tonnes": float,
        "total_co2e_tonnes":  float,
        "sources": { "DIESEL": float, "ELECTRICITY": float, ... }
      },
      "data_quality": {
        "overall_score": float,
        "overall_status": str,
        "actual_months":    int,
        "estimated_months": int,
        "missing_months":   int,
        "facilities": [ { facility_name, readiness_score, energy, water } ]
      },
      "generated_at": str,
    }
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _build_cover(wb, data)
    _build_energy(wb, data)
    _build_water(wb, data)
    _build_emissions(wb, data)
    _build_water_quality(wb, data)
    _build_data_quality(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Sheet 1 — Cover
# ---------------------------------------------------------------------------

def _build_cover(wb, data):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 4

    # Title banner
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 20

    ws.merge_cells("B3:D3")
    cell = ws["B3"]
    cell.value = "SANGATNA ESG Platform"
    cell.font = _font(bold=True, size=20, color=C_WHITE)
    cell.fill = _fill(C_GREEN_DARK)
    cell.alignment = _center()

    ws.merge_cells("B4:D4")
    cell = ws["B4"]
    cell.value = "Business Responsibility & Sustainability Report — Environmental Section"
    cell.font = _font(bold=False, size=11, color=C_WHITE)
    cell.fill = _fill(C_GREEN_MID)
    cell.alignment = _center()

    # Company details
    company = data.get("company", {})
    fy      = company.get("financial_year", "2025-26")

    details = [
        ("Company Name",     company.get("name", "—")),
        ("Industry",         company.get("industry", "—")),
        ("Location",         f"{company.get('city', '')} — {company.get('state', '')}"),
        ("GSTIN",            company.get("gstin", "—")),
        ("Financial Year",   f"FY {fy} (April 2025 – March 2026)"),
        ("Report Generated", data.get("generated_at",
                                      datetime.now().strftime("%d %b %Y %H:%M"))),
        ("Prepared by",      "SANGATNA AI-powered ESG Platform"),
    ]

    row = 6
    for label, value in details:
        ws.row_dimensions[row].height = 22
        _header_cell(ws, row, 2, label, bg=C_GREEN_LIGHT,
                     fg=C_GREEN_DARK, bold=True)
        _data_cell(ws, row, 3, value, bold=True if label == "Company Name" else False)
        row += 1

    # Scope summary boxes
    row += 2
    ws.merge_cells(
        start_row=row, start_column=2, end_row=row, end_column=4
    )
    _section_title(ws, row, 2, "Environmental Summary", width=3)
    row += 1

    energy   = data.get("energy", {})
    water    = data.get("water", {})
    emissions= data.get("emissions", {})

    summary = [
        ("Total Scope 1 Emissions",
         f"{emissions.get('scope1_co2e_tonnes', 0):.3f} tonnes CO₂e",
         C_BLUE_LIGHT),
        ("Total Scope 2 Emissions",
         f"{emissions.get('scope2_co2e_tonnes', 0):.3f} tonnes CO₂e",
         C_BLUE_LIGHT),
        ("Total GHG Emissions",
         f"{emissions.get('total_co2e_tonnes', 0):.3f} tonnes CO₂e",
         C_GREEN_LIGHT),
        ("Total Electricity Consumed",
         f"{energy.get('total_kwh', 0):,.0f} kWh",
         C_GRAY_LIGHT),
        ("Total Water Withdrawal",
         f"{water.get('total_withdrawal_kl', 0):,.2f} KL",
         C_BLUE_LIGHT),
        ("Water Recycled",
         f"{water.get('total_recycled_kl', 0):,.2f} KL",
         C_GREEN_LIGHT),
    ]

    for label, value, bg in summary:
        ws.row_dimensions[row].height = 22
        _data_cell(ws, row, 2, label, bg=bg, bold=False)
        _data_cell(ws, row, 3, value, bg=bg, bold=True, align="right")
        row += 1

    # Disclaimer
    row += 2
    ws.merge_cells(
        start_row=row, start_column=2, end_row=row, end_column=4
    )
    cell = ws.cell(
        row=row, column=2,
        value=(
            "This report is generated by SANGATNA AI-powered ESG platform. "
            "Data marked as ESTIMATED is derived from production intensity ratios. "
            "Actual figures should be verified against utility bills before submission."
        )
    )
    cell.font = _font(size=9, color=C_GRAY_MID, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40


# ---------------------------------------------------------------------------
# Sheet 2 — Energy
# ---------------------------------------------------------------------------

def _build_energy(wb, data):
    ws = wb.create_sheet("Energy")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 14   # Period
    ws.column_dimensions["B"].width = 22   # Facility
    ws.column_dimensions["C"].width = 18   # Source
    ws.column_dimensions["D"].width = 14   # kWh
    ws.column_dimensions["E"].width = 14   # Diesel (L)
    ws.column_dimensions["F"].width = 16   # CO₂e (tonnes)
    ws.column_dimensions["G"].width = 12   # Data type

    row = 1
    ws.merge_cells("A1:G1")
    _section_title(ws, row, 1, "Energy Consumption — FY " +
                   data.get("company", {}).get("financial_year", "2025-26"),
                   width=7)
    row += 1

    # Summary boxes
    energy = data.get("energy", {})
    summary_labels = [
        ("Total Electricity", f"{energy.get('total_kwh', 0):,.0f} kWh"),
        ("Total Diesel",      f"{energy.get('total_diesel_litres', 0):,.0f} L"),
        ("Scope 1 Emissions", f"{energy.get('scope1_co2e_tonnes', 0):.3f} tCO₂e"),
        ("Scope 2 Emissions", f"{energy.get('scope2_co2e_tonnes', 0):.3f} tCO₂e"),
        ("Total Emissions",   f"{energy.get('total_co2e_tonnes', 0):.3f} tCO₂e"),
    ]
    for i, (label, val) in enumerate(summary_labels):
        col = i + 1
        ws.row_dimensions[row].height = 18
        _data_cell(ws, row, col, label, bg=C_GREEN_LIGHT,
                   bold=True, color=C_GREEN_DARK)
        ws.row_dimensions[row + 1].height = 22
        _data_cell(ws, row + 1, col, val,
                   bold=True, align="right", color=C_GREEN_DARK)
    row += 3

    # Table headers
    headers = [
        "Period", "Facility", "Energy Source",
        "Electricity (kWh)", "Diesel (Litres)",
        "Emissions (tCO₂e)", "Data Type"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, row, col, h)
    row += 1

    # Monthly data
    monthly = energy.get("monthly", [])
    for i, m in enumerate(monthly):
        bg = C_AMBER_LIGHT if m.get("is_estimated") else (
            C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        )
        _data_cell(ws, row, 1, m.get("period", ""),       bg=bg)
        _data_cell(ws, row, 2, m.get("facility_name", ""), bg=bg)
        _data_cell(ws, row, 3, m.get("source", ""),        bg=bg)
        _data_cell(ws, row, 4, m.get("kwh"),
                   bg=bg, align="right", number_format="#,##0.00")
        _data_cell(ws, row, 5, m.get("diesel_litres"),
                   bg=bg, align="right", number_format="#,##0.00")
        _data_cell(ws, row, 6, m.get("co2e_tonnes"),
                   bg=bg, align="right", number_format="0.0000")
        dtype = "ESTIMATED" if m.get("is_estimated") else "ACTUAL"
        _status_cell(ws, row, 7, dtype)
        row += 1

    # Totals row
    _header_cell(ws, row, 1, "TOTAL", bg=C_GREEN_DARK)
    _header_cell(ws, row, 2, "", bg=C_GREEN_DARK)
    _header_cell(ws, row, 3, "", bg=C_GREEN_DARK)
    _header_cell(ws, row, 4,
                 f"{energy.get('total_kwh', 0):,.0f}",
                 bg=C_GREEN_DARK, align="center")
    _header_cell(ws, row, 5,
                 f"{energy.get('total_diesel_litres', 0):,.0f}",
                 bg=C_GREEN_DARK, align="center")
    _header_cell(ws, row, 6,
                 f"{energy.get('total_co2e_tonnes', 0):.4f}",
                 bg=C_GREEN_DARK, align="center")
    _header_cell(ws, row, 7, "", bg=C_GREEN_DARK)


# ---------------------------------------------------------------------------
# Sheet 3 — Water
# ---------------------------------------------------------------------------

def _build_water(wb, data):
    ws = wb.create_sheet("Water")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    water = data.get("water", {})
    fy    = data.get("company", {}).get("financial_year", "2025-26")

    row = 1
    ws.merge_cells("A1:F1")
    _section_title(ws, row, 1, f"Water Consumption — FY {fy}", width=6)
    row += 1

    # Summary
    summary = [
        ("Total Withdrawal",  f"{water.get('total_withdrawal_kl',  0):,.2f} KL"),
        ("Total Consumption", f"{water.get('total_consumption_kl', 0):,.2f} KL"),
        ("Total Recycled",    f"{water.get('total_recycled_kl',    0):,.2f} KL"),
        ("Total Discharged",  f"{water.get('total_discharged_kl',  0):,.2f} KL"),
    ]
    for i, (label, val) in enumerate(summary):
        col = i + 1
        _data_cell(ws, row,     col, label, bg=C_BLUE_LIGHT,
                   bold=True, color=C_BLUE)
        _data_cell(ws, row + 1, col, val,
                   bold=True, align="right", color=C_BLUE)
    row += 3

    # Headers
    headers = [
        "Period", "Facility", "Water Source",
        "Category", "Quantity (KL)", "Data Type"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, row, col, h)
    row += 1

    monthly = water.get("monthly", [])
    for i, m in enumerate(monthly):
        is_est = m.get("is_estimated") or \
                 m.get("entry_method") == "ESTIMATED"
        bg = C_AMBER_LIGHT if is_est else (
            C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        )
        _data_cell(ws, row, 1, m.get("period", ""),        bg=bg)
        _data_cell(ws, row, 2, m.get("facility_name", ""), bg=bg)
        _data_cell(ws, row, 3, m.get("source", ""),        bg=bg)
        _data_cell(ws, row, 4, m.get("category", ""),      bg=bg)
        _data_cell(ws, row, 5, m.get("quantity_kl"),
                   bg=bg, align="right", number_format="#,##0.0000")
        dtype = "ESTIMATED" if is_est else "ACTUAL"
        _status_cell(ws, row, 6, dtype)
        row += 1

    # Totals
    _header_cell(ws, row, 1, "TOTAL", bg=C_GREEN_DARK)
    for c in range(2, 5):
        _header_cell(ws, row, c, "", bg=C_GREEN_DARK)
    _header_cell(
        ws, row, 5,
        f"{water.get('total_withdrawal_kl', 0):,.2f}",
        bg=C_GREEN_DARK, align="center"
    )
    _header_cell(ws, row, 6, "", bg=C_GREEN_DARK)


# ---------------------------------------------------------------------------
# Sheet 4 — Emissions
# ---------------------------------------------------------------------------

def _build_emissions(wb, data):
    ws = wb.create_sheet("Emissions")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    emissions = data.get("emissions", {})
    fy        = data.get("company", {}).get("financial_year", "2025-26")

    row = 1
    ws.merge_cells("A1:D1")
    _section_title(ws, row, 1, f"GHG Emissions Summary — FY {fy}", width=4)
    row += 2

    # Scope summary
    scope_data = [
        ("Scope 1 — Direct Emissions",
         "Fuel combustion, diesel generators",
         emissions.get("scope1_co2e_tonnes", 0)),
        ("Scope 2 — Indirect Emissions",
         "Purchased electricity",
         emissions.get("scope2_co2e_tonnes", 0)),
        ("Total GHG Emissions",
         "Scope 1 + Scope 2",
         emissions.get("total_co2e_tonnes", 0)),
    ]

    _header_cell(ws, row, 1, "Emission Category")
    _header_cell(ws, row, 2, "Description")
    _header_cell(ws, row, 3, "CO₂e (tonnes)", align="center")
    _header_cell(ws, row, 4, "CO₂e (kg)",     align="center")
    row += 1

    for i, (scope, desc, val) in enumerate(scope_data):
        bg = C_GREEN_LIGHT if i == 2 else (
            C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        )
        bold = i == 2
        _data_cell(ws, row, 1, scope, bg=bg, bold=bold)
        _data_cell(ws, row, 2, desc,  bg=bg)
        _data_cell(ws, row, 3, round(val, 4),
                   bg=bg, align="right",
                   number_format="0.0000", bold=bold)
        _data_cell(ws, row, 4, round(val * 1000, 1),
                   bg=bg, align="right",
                   number_format="#,##0.0", bold=bold)
        row += 1

    # Source breakdown
    row += 1
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=4
    )
    _section_title(ws, row, 1, "Emissions by Source", width=4)
    row += 1

    _header_cell(ws, row, 1, "Emission Source")
    _header_cell(ws, row, 2, "Scope")
    _header_cell(ws, row, 3, "CO₂e (tonnes)", align="center")
    _header_cell(ws, row, 4, "% of Total",    align="center")
    row += 1

    sources = emissions.get("sources", {})
    total   = emissions.get("total_co2e_tonnes", 0) or 1
    scope_map = {
        "ELECTRICITY": "Scope 2",
        "DIESEL":      "Scope 1",
        "LPG":         "Scope 1",
        "CNG":         "Scope 1",
        "FURNACE_OIL": "Scope 1",
    }

    for i, (source, val) in enumerate(sources.items()):
        bg  = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        pct = round((val / total) * 100, 1) if total else 0
        _data_cell(ws, row, 1, source.replace("_", " "), bg=bg)
        _data_cell(ws, row, 2, scope_map.get(source, "Scope 1"), bg=bg)
        _data_cell(ws, row, 3, round(val, 4),
                   bg=bg, align="right", number_format="0.0000")
        _data_cell(ws, row, 4, f"{pct}%",
                   bg=bg, align="right")
        row += 1

    # Emission factors note
    row += 1
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=4
    )
    cell = ws.cell(
        row=row, column=1,
        value=(
            "Emission Factors: Electricity — CEA CO2 Baseline Database "
            "(state-wise grid factors). Diesel — IPCC AR6 (2.68 kg CO₂e/litre). "
            "LPG — IPCC AR6 (2.98 kg CO₂e/kg)."
        )
    )
    cell.font = _font(size=9, italic=True, color=C_GRAY_MID)
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30


# ---------------------------------------------------------------------------
# Sheet 5 — Water Quality
# ---------------------------------------------------------------------------

def _build_water_quality(wb, data):
    ws = wb.create_sheet("Water Quality")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    fy = data.get("company", {}).get("financial_year", "2025-26")
    row = 1
    ws.merge_cells("A1:G1")
    _section_title(ws, row, 1,
                   f"Water Quality & Compliance — FY {fy}", width=7)
    row += 2

    samples = data.get("water_quality", [])
    if not samples:
        ws.cell(row=row, column=1,
                value="No water quality samples uploaded for this period.")
        return

    for sample in samples:
        # Sample header
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=7
        )
        cell = ws.cell(
            row=row, column=1,
            value=(
                f"Lab: {sample.get('lab_name', '—')}  |  "
                f"Date: {sample.get('collection_date', '—')}  |  "
                f"Type: {sample.get('water_type', '—')}  |  "
                f"Overall: {sample.get('overall_status', '—')}"
            )
        )
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_PURPLE)
        cell.alignment = _left()
        row += 1

        # Column headers
        headers = [
            "Parameter", "Category", "Measured Value",
            "Unit", "BIS Limit", "CPCB Limit", "Status"
        ]
        for col, h in enumerate(headers, 1):
            _header_cell(ws, row, col, h, bg=C_PURPLE_LIGHT,
                         fg=C_PURPLE, bold=True)
        row += 1

        # Readings
        for i, reading in enumerate(sample.get("readings", [])):
            bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
            _data_cell(ws, row, 1, reading.get("parameter", ""),  bg=bg, bold=True)
            _data_cell(ws, row, 2, reading.get("category", ""),   bg=bg)
            _data_cell(ws, row, 3, reading.get("measured_value"), bg=bg,
                       align="right", number_format="0.000000")
            _data_cell(ws, row, 4, reading.get("unit", ""),       bg=bg)
            _data_cell(ws, row, 5, reading.get("bis_limit"),      bg=bg,
                       align="right")
            _data_cell(ws, row, 6, reading.get("cpcb_limit"),     bg=bg,
                       align="right")
            status = reading.get("status", "NO_DATA")
            _status_cell(ws, row, 7, status)
            row += 1

        row += 1  # Gap between samples


# ---------------------------------------------------------------------------
# Sheet 6 — Data Quality
# ---------------------------------------------------------------------------

def _build_data_quality(wb, data):
    ws = wb.create_sheet("Data Quality")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    fy  = data.get("company", {}).get("financial_year", "2025-26")
    dq  = data.get("data_quality", {})
    row = 1

    ws.merge_cells("A1:F1")
    _section_title(ws, row, 1,
                   f"Data Quality & Completeness — FY {fy}", width=6)
    row += 2

    # Overall score
    score  = dq.get("overall_score", 0)
    status = dq.get("overall_status", "INSUFFICIENT")
    score_bg = (
        C_GREEN_LIGHT if score >= 90 else
        C_AMBER_LIGHT if score >= 60 else
        C_RED_LIGHT
    )
    score_fg = (
        C_GREEN_DARK if score >= 90 else
        C_AMBER      if score >= 60 else
        C_RED
    )

    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=3
    )
    cell = ws.cell(row=row, column=1,
                   value=f"Overall Readiness Score: {score}/100 — {status}")
    cell.font  = _font(bold=True, size=13, color=score_fg)
    cell.fill  = _fill(score_bg)
    cell.alignment = _center()
    ws.row_dimensions[row].height = 28
    row += 2

    # Summary counts
    summary = [
        ("Actual data months",    dq.get("actual_months",    0), C_GREEN_LIGHT),
        ("Estimated data months", dq.get("estimated_months", 0), C_AMBER_LIGHT),
        ("Missing data months",   dq.get("missing_months",   0), C_RED_LIGHT),
    ]
    for label, val, bg in summary:
        _data_cell(ws, row, 1, label, bg=bg, bold=True)
        _data_cell(ws, row, 2, val,   bg=bg,
                   bold=True, align="right")
        row += 1

    row += 1

    # Per-facility breakdown
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=6
    )
    _section_title(ws, row, 1, "Facility Breakdown", width=6)
    row += 1

    headers = [
        "Facility", "Readiness Score",
        "Energy Actual", "Energy Est.", "Water Actual", "Water Est."
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, row, col, h)
    row += 1

    for i, fac in enumerate(dq.get("facilities", [])):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        fac_score = fac.get("readiness_score", 0)
        score_col = (
            C_GREEN_DARK if fac_score >= 90 else
            C_AMBER      if fac_score >= 60 else
            C_RED
        )
        _data_cell(ws, row, 1, fac.get("facility_name", ""), bg=bg, bold=True)
        _data_cell(ws, row, 2, f"{fac_score}/100",
                   bg=bg, bold=True, color=score_col, align="right")
        energy = fac.get("energy", {})
        water  = fac.get("water",  {})
        _data_cell(ws, row, 3, energy.get("actual_months",    0),
                   bg=bg, align="right")
        _data_cell(ws, row, 4, energy.get("estimated_months", 0),
                   bg=bg, align="right")
        _data_cell(ws, row, 5, water.get("actual_months",     0),
                   bg=bg, align="right")
        _data_cell(ws, row, 6, water.get("estimated_months",  0),
                   bg=bg, align="right")
        row += 1

    # Legend
    row += 2
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=6
    )
    _section_title(ws, row, 1, "Legend", width=6)
    row += 1

    legend = [
        ("ACTUAL",    "Data from real utility bills uploaded to SANGATNA",
         C_GREEN_LIGHT, C_GREEN_DARK),
        ("ESTIMATED", "Data estimated from production intensity ratios",
         C_AMBER_LIGHT, C_AMBER),
        ("MISSING",   "No data available — bill not uploaded, cannot estimate",
         C_RED_LIGHT,  C_RED),
    ]
    for label, desc, bg, fg in legend:
        ws.row_dimensions[row].height = 20
        cell = ws.cell(row=row, column=1, value=label)
        cell.font      = _font(bold=True, color=fg)
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.alignment = _center()
        _data_cell(ws, row, 2, desc, bg=bg)
        ws.merge_cells(
            start_row=row, start_column=2,
            end_row=row,   end_column=6
        )
        row += 1
